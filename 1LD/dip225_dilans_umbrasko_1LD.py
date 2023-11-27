from openpyxl import load_workbook
wb = load_workbook('1LD/test1.xlsx')
ws = wb.active
max_row = ws.max_row
total=0
over3k=0
for row in range(2, max_row+1):
    hour = ws['B' + str(row)].value
    rate = ws['C' + str(row)].value
    if type(hour) in (int, float) and type(rate) in (int, float):
        salary = hour*rate
        total += salary
        if salary > 3000:
            over3k += 1
print(over3k)

wb.save('result1.xlsx')
wb.close()